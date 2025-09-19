#!/usr/bin/env python3
"""
excel_keyinfo_extractor.py

Usage:
    python excel_keyinfo_extractor.py --input INPUT.xlsx --out_json output.json --tables_dir tables_out/

This script scans all sheets of an Excel workbook and extracts two kinds of content:
1) Key-value pairs (simple two-cell rows, where the first non-empty cell is a text key)
2) Tables (header-like row followed by data rows)
3) Structured information from specific sheets (General Info, Energy-Waste Info, Recommendation Info)

It writes a JSON summary and exports each detected table as a CSV file.
"""

import argparse
import json
import logging
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def is_textish(x) -> bool:
    """Check if a value is text-like and suitable as a key."""
    if x is None:
        return False
    if isinstance(x, str):
        s = x.strip()
        return len(s) > 0 and len(s) <= 200 and not s.replace('.', '').replace('-', '').isdigit()
    return False

def safe_convert_numeric(value: Any) -> Union[float, int, str, None]:
    """Safely convert value to numeric type if possible, otherwise return as string."""
    if value is None or (isinstance(value, str) and value.strip() == ''):
        return None
    
    try:
        # Try to convert to float first
        float_val = float(value)
        # If it's a whole number, return as int
        if float_val.is_integer():
            return int(float_val)
        return float_val
    except (ValueError, TypeError):
        # Return as cleaned string if conversion fails
        return str(value).strip() if value is not None else None

def clean_key(key: str) -> str:
    """Clean and standardize key names."""
    if not key:
        return key
    # Remove colons, extra whitespace, and normalize
    cleaned = re.sub(r'[:]+', '', str(key)).strip()
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned

def nonempty_count(cells: List[Any]) -> int:
    """Count non-empty cells in a list."""
    return sum(1 for c in cells if c not in (None, "", " "))

def row_values(ws: Worksheet, row_idx: int, max_col: int) -> List[Any]:
    """Extract values from a row in a worksheet."""
    try:
        return [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
    except Exception as e:
        logging.warning(f"Error reading row {row_idx}: {e}")
        return [None] * max_col

def detect_key_value_pairs(ws: Worksheet, max_col: int = 12) -> List[Dict[str, Any]]:
    """Detect key-value pairs in a worksheet with improved robustness."""
    pairs = []
    
    for r in range(1, ws.max_row + 1):
        vals = row_values(ws, r, max_col)
        cnt = nonempty_count(vals)
        
        # Allow 2-3 non-empty cells for key-value pairs (some may have validation info)
        if 2 <= cnt <= 3:
            idxs = [i for i, v in enumerate(vals) if v not in (None, "", " ")]
            
            if len(idxs) >= 2:
                k, v = vals[idxs[0]], vals[idxs[1]]
                if is_textish(k):
                    cleaned_key = clean_key(str(k))
                    cleaned_value = safe_convert_numeric(v)
                    
                    if cleaned_key:  # Only add if key is meaningful
                        pairs.append({
                            "row": r, 
                            "key": cleaned_key, 
                            "value": cleaned_value,
                            "raw_key": str(k).strip() if k else "",
                            "raw_value": v
                        })
    
    # Dedup by cleaned key, keeping the first occurrence
    seen = set()
    deduped = []
    for item in pairs:
        key_lower = item["key"].lower()
        if key_lower not in seen:
            seen.add(key_lower)
            deduped.append(item)
    
    return deduped

def looks_like_header(row: List[Any]) -> bool:
    """Check if a row looks like a table header with improved detection."""
    vals = [v for v in row if v not in (None, "", " ")]
    if len(vals) < 2:  # Reduced minimum for IAC templates
        return False
    
    # Check for string ratio
    str_count = sum(1 for v in vals if isinstance(v, str) and len(str(v).strip()) > 0)
    str_ratio = str_count / len(vals)
    
    # Check for typical header patterns
    header_indicators = ['info', 'code', 'name', 'description', 'unit', 'cost', 'consumption', 'savings']
    has_header_words = any(any(indicator in str(v).lower() for indicator in header_indicators) for v in vals if v is not None)
    
    return str_ratio >= 0.4 or has_header_words

def extend_table_down(ws: Worksheet, start_row: int, max_col: int) -> int:
    r = start_row + 1
    blank_run = 0
    while r <= ws.max_row:
        vals = row_values(ws, r, max_col)
        if nonempty_count(vals) == 0:
            blank_run += 1
            if blank_run >= 2:
                return r - 1
        else:
            blank_run = 0
        r += 1
    return ws.max_row

def clean_headers(headers: List[Any]) -> List[str]:
    cleaned = []
    used = set()
    for i, h in enumerate(headers, 1):
        if h is None or str(h).strip() == "":
            name = f"col_{i}"
        else:
            name = str(h).strip()
        base = name
        j = 1
        while name in used:
            name = f"{base}_{j}"
            j += 1
        used.add(name)
        cleaned.append(name)
    return cleaned

def detect_tables(ws: Worksheet, max_col: int = 50) -> List[Dict[str, Any]]:
    tables = []
    r = 1
    while r <= ws.max_row:
        row = row_values(ws, r, max_col)
        if looks_like_header(row):
            if r + 1 <= ws.max_row and nonempty_count(row_values(ws, r + 1, max_col)) >= 1:
                end_r = extend_table_down(ws, r, max_col)
                headers = clean_headers(row)
                data = []
                for rr in range(r + 1, end_r + 1):
                    vals = row_values(ws, rr, max_col)
                    if nonempty_count(vals) == 0:
                        continue
                    row_dict = {headers[i]: vals[i] for i in range(min(len(headers), len(vals)))}
                    data.append(row_dict)
                if data:
                    tables.append({
                        "start_row": r,
                        "end_row": end_r,
                        "headers": headers,
                        "rows": data,
                    })
                r = end_r + 1
                continue
        r += 1
    return tables

def export_tables_to_csv(sheet_name: str, tables: List[Dict[str, Any]], out_dir: Path) -> List[str]:
    out_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    for idx, t in enumerate(tables, 1):
        df = pd.DataFrame(t["rows"])
        safe_sheet = "".join(ch if ch.isalnum() or ch in "-_." else "_" for ch in sheet_name)[:40]
        fname = out_dir / f"{safe_sheet}_table{idx}.csv"
        df.to_csv(fname, index=False)
        paths.append(str(fname))
    return paths

def extract_excel_key_info(xlsx_path: str, tables_dir: Path) -> Dict[str, Any]:
    """Extract key information from Excel file with comprehensive error handling."""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        result = {
            "workbook": Path(xlsx_path).name, 
            "file_path": str(Path(xlsx_path).resolve()),
            "sheets": {},
            "extraction_metadata": {
                "total_sheets": len(wb.worksheets),
                "sheet_names": [ws.title for ws in wb.worksheets]
            }
        }

        for ws in wb.worksheets:
            try:
                logging.info(f"Processing sheet: {ws.title}")
                
                # Determine reasonable max columns based on actual data
                actual_max_col = min(ws.max_column or 50, 100) if ws.max_column else 50
                
                kvs = detect_key_value_pairs(ws, max_col=min(actual_max_col, 20))
                tables = detect_tables(ws, max_col=actual_max_col)
                csv_files = export_tables_to_csv(ws.title, tables, tables_dir)
                
                result["sheets"][ws.title] = {
                    "key_values": kvs,
                    "table_count": len(tables),
                    "table_csv_files": csv_files,
                    "dimensions": {
                        "max_row": ws.max_row,
                        "max_column": ws.max_column
                    },
                    "key_value_count": len(kvs)
                }
                
                logging.info(f"Sheet '{ws.title}': {len(kvs)} key-value pairs, {len(tables)} tables")
                
            except Exception as e:
                logging.error(f"Error processing sheet '{ws.title}': {e}")
                result["sheets"][ws.title] = {
                    "error": str(e),
                    "key_values": [],
                    "table_count": 0,
                    "table_csv_files": []
                }
                
        return result
        
    except InvalidFileException as e:
        logging.error(f"Invalid Excel file {xlsx_path}: {e}")
        raise
    except FileNotFoundError as e:
        logging.error(f"File not found {xlsx_path}: {e}")
        raise
    except Exception as e:
        logging.error(f"Unexpected error processing {xlsx_path}: {e}")
        raise

# New specialized extraction functions for IAC Assessment Templates

def extract_general_info_dict(xlsx_path: str) -> Dict[str, Any]:
    """Extract General Info sheet as a structured dictionary with standardized keys."""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        
        if 'General Info' not in [ws.title for ws in wb.worksheets]:
            logging.warning("General Info sheet not found")
            return {}
            
        ws = wb['General Info']
        info_dict = {}
        
        # Field mapping to match document_extractor.py standardized keys
        field_mapping = {
            'SIC Code (4 Digits)': 'sic_no',
            'SIC Code: (4 Digits)': 'sic_no',
            'SIC. No.': 'sic_no',
            'SIC No.': 'sic_no',
            'SIC No': 'sic_no',
            'NAICS Code (6 Digits)': 'naics_code',
            'NAICS Code: (6 Digits)': 'naics_code',
            'NAICS Code': 'naics_code',
            'Principal Product': 'principal_product',
            'Principle Product': 'principal_product',  # Handle typo in Excel
            'Principal Products': 'principal_products',
            '# of Employees': 'no_of_employees',
            'No. of Employees': 'no_of_employees',
            'Number of Employees': 'no_of_employees',
            'Plant Area (sqft.)': 'total_facility_area',
            'Total Facility Area': 'total_facility_area',
            'Production Hrs. Annual': 'operating_hours',
            'Operating Hours': 'operating_hours',
            'Annual Production': 'annual_production',
            'Annual Sales ($)': 'annual_sales',
            'Annual Sales': 'annual_sales',
            'Production Units': 'production_units',
            'Motor Horsepower Capacity': 'motor_horsepower_capacity',
            'Largest Motor Horsepower': 'largest_motor_horsepower',
            'Steam Capacity(LBM/Hr)': 'steam_capacity_lbm_hr',
            'Max Steam Pressure(PSIG)': 'max_steam_pressure_psig',
            'Air Compressor HP': 'air_compressor_hp',
            'Max Compressed Air Press(PSIG)': 'max_compressed_air_press_psig',
            'Value per Finished Product': 'value_per_finished_product',
            'Total Energy Usage': 'total_energy_usage',
            'Total Utility Cost': 'total_utility_cost',
            'No. of Assessment Recommendations': 'no_of_assessment_recommendations'
        }
        
        # Extract key-value pairs from first two columns
        for r in range(1, ws.max_row + 1):
            key_cell = ws.cell(row=r, column=1).value
            value_cell = ws.cell(row=r, column=2).value
            
            if key_cell and str(key_cell).strip():
                raw_key = str(key_cell).strip()
                cleaned_key = clean_key(raw_key)
                cleaned_value = safe_convert_numeric(value_cell)
                
                # Map to standardized key name
                standardized_key = field_mapping.get(raw_key, field_mapping.get(cleaned_key, None))
                
                if standardized_key:
                    # Keep string values for product fields, convert others to numeric
                    if standardized_key in ['principal_product', 'principal_products']:
                        info_dict[standardized_key] = str(value_cell).strip() if value_cell else ""
                    else:
                        info_dict[standardized_key] = cleaned_value
                elif cleaned_key and cleaned_key not in ['GENERAL INFO']:  # Skip header rows
                    # If no mapping found, use cleaned key as fallback
                    fallback_key = cleaned_key.lower().replace(' ', '_').replace('.', '').replace('(', '').replace(')', '').replace(':', '')
                    if standardized_key in ['principal_product', 'principal_products'] or 'product' in fallback_key:
                        info_dict[fallback_key] = str(value_cell).strip() if value_cell else ""
                    else:
                        info_dict[fallback_key] = cleaned_value
        
        return info_dict
        
    except Exception as e:
        logging.error(f"Error extracting General Info: {e}")
        return {}

def extract_energy_waste_info_dict(xlsx_path: str) -> Dict[str, Any]:
    """Extract Energy-Waste Info sheet as a structured dictionary with standardized energy type names."""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        
        if 'Energy-Waste Info' not in [ws.title for ws in wb.worksheets]:
            logging.warning("Energy-Waste Info sheet not found")
            return {}
            
        ws = wb['Energy-Waste Info']
        energy_info = {
            "data": [],  # Changed from "energy_sources" to match document_extractor format
            "period": {"start": "", "end": ""},  # Added period info
            "summary": {}
        }
        
        # Energy type mapping to match document_extractor.py standardized names
        energy_type_mapping = {
            'Electrical Consumption': 'electrical_energy',
            'Electric Consumption': 'electrical_energy', 
            'Electrical Energy': 'electrical_energy',
            'Electric Energy': 'electrical_energy',
            'Electricity': 'electrical_energy',
            'Electrical Demand': 'electrical_demand',
            'Electric Demand': 'electrical_demand',
            'Demand Charge': 'demand_charge',
            'Demand': 'electrical_demand',
            'Other Electrical Fees': 'electrical_fees',
            'Electrical Fees': 'electrical_fees',
            'Natural Gas': 'natural_gas',
            'L.P.G.': 'propane_gas',
            'LPG': 'propane_gas',
            'Propane': 'propane_gas',
            'Propane Gas': 'propane_gas',
            '#1 Fuel Oil': 'fuel_oil',
            '#2 Fuel Oil': 'fuel_oil',
            '#4 Fuel Oil': 'fuel_oil', 
            '#6 Fuel Oil': 'fuel_oil',
            'Fuel Oil #1': 'fuel_oil',
            'Fuel Oil #2': 'fuel_oil',
            'Fuel Oil #4': 'fuel_oil',
            'Fuel Oil #6': 'fuel_oil',
            'Fuel Oil': 'fuel_oil',
            'Heating Oil': 'heating_oil',
            'Coal': 'coal',
            'Wood': 'biomass',
            'Paper': 'biomass',
            'Other Gas': 'other_gas',
            'Other Energy': 'other_energy',
            'Steam': 'steam',
            'Water Usage': 'water',
            'Water Disposal': 'water_disposal',
            'Other Liquid (non-haz)': 'other_liquid_non_haz',
            'Other Liquid (haz)': 'other_liquid_haz',
            'Solid Waste (non-haz)': 'solid_waste_non_haz',
            'Solid Waste (haz)': 'solid_waste_haz',
            'Gaseous Waste': 'gaseous_waste',
            'Compressed Air': 'compressed_air',
            'Chilled Water': 'chilled_water',
            'Hot Water': 'hot_water',
            'Total Utility': 'total_utility',
            'Total': 'total_utility'
        }
        
        # Find the data table (usually starts around row 3)
        tables = detect_tables(ws, max_col=min(ws.max_column or 15, 15))
        
        if tables:
            main_table = tables[0]  # Use the first/main table
            
            total_electrical_cost = 0
            total_energy_cost = 0
            total_utility_cost = 0
            
            for row in main_table["rows"]:
                source_name = None
                consumption = None
                cost = None
                unit_cost = None
                consumption_units = ""
                total_cost = None
                
                # Find relevant columns with improved column mapping
                for key, value in row.items():
                    key_lower = str(key).lower()
                    
                    if 'energy' in key_lower or 'waste' in key_lower or 'info' in key_lower:
                        if value and str(value).strip():
                            source_name = str(value).strip()
                    elif key_lower == 'consumption_1':  # This has the actual consumption values
                        consumption = safe_convert_numeric(value)
                    elif key_lower == 'cost_1':  # This has the actual cost values
                        cost = safe_convert_numeric(value)
                    elif key_lower == 'cost' and not cost:  # Fallback to original cost column
                        cost = safe_convert_numeric(value)
                    elif key_lower in ['col_5']:  # This column contains the units (kWh, MMBtu, etc.)
                        if value and str(value).strip() and str(value).strip() not in ['', 'n/a']:
                            consumption_units = str(value).strip()
                    elif 'unit' in key_lower and 'cost' in key_lower:
                        unit_cost = safe_convert_numeric(value)
                    elif key_lower in ['col_12']:  # Sometimes total costs appear in rightmost columns
                        if value and isinstance(value, (int, float)) and value > 0:
                            total_cost = safe_convert_numeric(value)
                
                if source_name and (consumption is not None or cost is not None):
                    # Map to standardized energy type
                    standardized_type = energy_type_mapping.get(source_name, 
                                                             source_name.lower()
                                                             .replace(' ', '_')
                                                             .replace('-', '_')
                                                             .replace('&', 'and')
                                                             .replace('/', '_')
                                                             .replace('(', '')
                                                             .replace(')', '')
                                                             .replace('.', ''))
                    
                    # Create usage data structure similar to document_extractor
                    usage_data = {}
                    if consumption is not None and consumption_units:
                        # Use proper unit format like "kWh/yr" for consistency
                        unit_key = consumption_units
                        if not unit_key.endswith('/yr') and unit_key not in ['kW months/yr']:
                            unit_key = f"{consumption_units}/yr" if consumption_units else "value"
                        usage_data[unit_key] = consumption
                    elif consumption is not None:
                        usage_data["value"] = consumption
                    
                    # Parse unit cost information
                    unit_cost_data = None
                    if unit_cost is not None and unit_cost != 0:
                        # Clean up unit for unit cost (remove "per " prefix)
                        clean_unit = consumption_units
                        if consumption_units:
                            clean_unit = consumption_units.replace('per ', '').strip()
                            unit_cost_data = {
                                "amount": unit_cost,
                                "unit": clean_unit
                            }
                        else:
                            unit_cost_data = {
                                "amount": unit_cost,
                                "unit": "unknown"
                            }
                    
                    # Use the higher total cost if available (for electrical fees, etc.)
                    final_cost = cost if cost is not None else 0
                    if total_cost and total_cost > final_cost:
                        final_cost = total_cost
                    
                    source_info = {
                        "type": standardized_type,
                        "original_name": source_name,  # Keep original for reference
                        "usage": usage_data,
                        "cost": final_cost,
                        "unit_cost": unit_cost_data
                    }
                    energy_info["data"].append(source_info)
                    
                    # Accumulate totals (only for positive numeric costs)
                    if final_cost and isinstance(final_cost, (int, float)) and final_cost > 0:
                        # Only add to totals if not a utility total (to avoid double counting)
                        if standardized_type != 'total_utility':
                            total_energy_cost += final_cost
                            
                            # Categorize costs
                            if standardized_type in ['electrical_energy', 'electrical_demand', 'electrical_fees']:
                                total_electrical_cost += final_cost
                        else:
                            total_utility_cost = final_cost  # Store but don't add to avoid double counting
            
            # Build summary matching document_extractor patterns
            energy_info["summary"] = {
                "total_energy_cost": total_energy_cost,
                "total_electrical_cost": total_electrical_cost,
                "total_utility_cost": total_utility_cost if total_utility_cost > 0 else total_energy_cost,
                "num_energy_sources": len([item for item in energy_info["data"] if item["cost"] > 0]),
                "total_data_entries": len(energy_info["data"])
            }
        
        return energy_info
        
    except Exception as e:
        logging.error(f"Error extracting Energy-Waste Info: {e}")
        return {}

def extract_recommendation_info_dict(xlsx_path: str) -> Dict[str, Any]:
    """Extract Recommendation Info sheet as a structured dictionary."""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        
        if 'Recommendation Info' not in [ws.title for ws in wb.worksheets]:
            logging.warning("Recommendation Info sheet not found")
            return {}
            
        ws = wb['Recommendation Info']
        reco_info = {
            "recommendations": [],
            "summary": {},
            "totals": {}
        }
        
        # Find the main recommendations table
        tables = detect_tables(ws, max_col=min(ws.max_column or 35, 35))
        
        if tables:
            main_table = tables[0]  # Use the first/main table
            
            total_savings = 0
            total_cost = 0
            valid_recommendations = 0
            
            for row in main_table["rows"]:
                rec_data = {}
                
                # Extract key fields with flexible mapping
                for key, value in row.items():
                    key_lower = str(key).lower()
                    
                    if value is not None and str(value).strip():
                        if 'arc' in key_lower and 'code' in key_lower:
                            rec_data['arc_code'] = safe_convert_numeric(value)
                        elif 'app' in key_lower and 'code' in key_lower:
                            rec_data['app_code'] = safe_convert_numeric(value)
                        elif 'description' in key_lower:
                            rec_data['description'] = str(value).strip()
                        elif 'primary' in key_lower and 'resource' in key_lower:
                            rec_data['primary_resource'] = str(value).strip()
                        elif 'unit' in key_lower and 'savings' in key_lower:
                            rec_data['unit_savings'] = safe_convert_numeric(value)
                        elif 'savings' in key_lower and '$' in key_lower:
                            rec_data['dollar_savings'] = safe_convert_numeric(value)
                        elif 'cost' in key_lower and 'capital' in key_lower:
                            rec_data['capital_cost'] = safe_convert_numeric(value)
                        elif 'cost' in key_lower and 'other' in key_lower:
                            rec_data['other_cost'] = safe_convert_numeric(value)
                
                # Only include if it has meaningful data
                if any(v is not None for v in rec_data.values()) and len(rec_data) > 1:
                    reco_info["recommendations"].append(rec_data)
                    valid_recommendations += 1
                    
                    # Accumulate totals
                    if 'dollar_savings' in rec_data and isinstance(rec_data['dollar_savings'], (int, float)):
                        total_savings += rec_data['dollar_savings']
                    
                    if 'capital_cost' in rec_data and isinstance(rec_data['capital_cost'], (int, float)):
                        total_cost += rec_data['capital_cost']
                    if 'other_cost' in rec_data and isinstance(rec_data['other_cost'], (int, float)):
                        total_cost += rec_data['other_cost']
            
            reco_info["summary"] = {
                "total_recommendations": valid_recommendations,
                "total_annual_savings": total_savings,
                "total_implementation_cost": total_cost,
                "simple_payback_years": total_savings / total_cost if total_cost > 0 else 0
            }
        
        return reco_info
        
    except Exception as e:
        logging.error(f"Error extracting Recommendation Info: {e}")
        return {}

def extract_all_structured_info(xlsx_path: str) -> Dict[str, Any]:
    """Extract all structured information from an IAC Assessment Template."""
    return {
        "general_info": extract_general_info_dict(xlsx_path),
        "energy_waste_info": extract_energy_waste_info_dict(xlsx_path),
        "recommendation_info": extract_recommendation_info_dict(xlsx_path),
        "extraction_timestamp": pd.Timestamp.now().isoformat(),
        "source_file": str(Path(xlsx_path).resolve())
    }

def main():
    ap = argparse.ArgumentParser(description="Extract key information from IAC Assessment Excel files")
    ap.add_argument("--input", required=True, help="Path to Excel file")
    ap.add_argument("--out_json", required=True, help="Path to write JSON summary")
    ap.add_argument("--tables_dir", default="tables_out", help="Folder to write CSV tables")
    ap.add_argument("--structured_only", action="store_true", 
                    help="Extract only structured info (General, Energy-Waste, Recommendation)")
    ap.add_argument("--verbose", "-v", action="store_true", help="Enable verbose logging")
    args = ap.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    try:
        tables_dir = Path(args.tables_dir)
        
        if args.structured_only:
            # Extract only the structured information
            summary = extract_all_structured_info(args.input)
        else:
            # Extract everything (original behavior + structured info)
            summary = extract_excel_key_info(args.input, tables_dir)
            # Add structured info to the summary
            summary["structured_extracts"] = extract_all_structured_info(args.input)

        # Write output
        with open(args.out_json, "w", encoding="utf-8") as f:
            json.dump(summary, f, indent=2, ensure_ascii=False, default=str)

        print(f"✓ Wrote JSON summary to: {args.out_json}")
        if not args.structured_only:
            print(f"✓ CSV tables to: {tables_dir.resolve()}")
        
        # Print summary statistics
        if "structured_extracts" in summary:
            structured = summary["structured_extracts"]
            print(f"\n📊 Extraction Summary:")
            if structured["general_info"]:
                print(f"   General Info: {len(structured['general_info'])} fields")
            if structured["energy_waste_info"].get("energy_sources"):
                print(f"   Energy Sources: {len(structured['energy_waste_info']['energy_sources'])}")
            if structured["recommendation_info"].get("recommendations"):
                print(f"   Recommendations: {len(structured['recommendation_info']['recommendations'])}")
                total_savings = structured["recommendation_info"].get("summary", {}).get("total_annual_savings", 0)
                if total_savings:
                    print(f"   Total Annual Savings: ${total_savings:,.2f}")
            
    except Exception as e:
        logging.error(f"Failed to process file: {e}")
        raise

if __name__ == "__main__":
    main()
